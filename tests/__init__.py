# -*- coding: utf-8 -*-
import io
import unittest
from datetime import date

from openpyxl.reader.excel import load_workbook

from xml2xlsx import xml2xlsx


class XML2XLSXTest(unittest.TestCase):
    def test_single_row(self):
        template = """
        <sheet name="test">
            <row>
                <cell>test cell</cell>
                <cell>test cell2</cell>
            </row>
        </sheet>

        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)

        self.assertEquals(len(wb.worksheets), 1,
                          u"Created workbook should have only one sheet")
        self.assertIn("test", wb.get_sheet_names(), u"Worksheet 'test' missing")
        ws = wb.get_sheet_by_name("test")
        self.assertEquals(ws["A1"].value, u"test cell")
        self.assertEquals(ws["B1"].value, u"test cell2")

    def test_unicode(self):
        template = """
        <sheet name="test">
            <row><cell>aąwźćńół</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb.get_sheet_by_name("test")
        self.assertEquals(ws["A1"].value, u"aąwźćńół")

    def test_multiple_rows(self):
        template = """
        <sheet name="test">
            <row>
                <cell>test cell</cell>
            </row>
            <row>
                <cell>test cell2</cell>
            </row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb.get_sheet_by_name("test")
        self.assertEquals(ws["A1"].value, u"test cell")
        self.assertEquals(ws["A2"].value, u"test cell2")

    def test_cell_type_number(self):
        template = u"""
        <sheet name="test"><row><cell type="number">1123.4</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb.get_sheet_by_name("test")
        self.assertEquals(ws["A1"].value, 1123.4)

    def test_cell_type_date(self):
        template = u"""
        <sheet name="test">
            <row><cell type="date" date-fmt="%d.%m.%Y">24.01.1981</cell></row>
        </sheet>
        """
        sheet = io.BytesIO(xml2xlsx(template))
        wb = load_workbook(sheet)
        ws = wb.get_sheet_by_name("test")
        self.assertEquals(ws["A1"].value.date(), date(1981, 01, 24))


if __name__ == '__main__':
    unittest.main()
